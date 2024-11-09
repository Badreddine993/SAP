from django.shortcuts import render ,get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import Category, Expense
from django.contrib import messages
from django.shortcuts import redirect 
from django.core.paginator import Paginator
import json
from django.http import HttpResponse
from django.http import JsonResponse
from userpreferences.models import UserPreferences
import datetime
from django.db.models import Sum
import csv
import openpyxl
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph 
from reportlab.lib.pagesizes import letter


#lets import request form django



# Create your views here.
@login_required(login_url='/auth/login')
def search_expenses(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        expenses = Expense.objects.filter(
            amount__istartswith=search_str, user=request.user
        ) | Expense.objects.filter(
            date__istartswith=search_str, user=request.user
        ) | Expense.objects.filter(
            description__icontains=search_str, user=request.user
        ) | Expense.objects.filter(
            category__icontains=search_str, user=request.user
        )
        data = expenses.values()
        return JsonResponse(list(data), safe=False)

def index(request):

    categories = Category.objects.all()
    #lets get the expenses off our user
    expenses = Expense.objects.filter(user=request.user)
    paginator = Paginator(expenses, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    try:
        currency = UserPreferences.objects.get(user=request.user).currency
    except UserPreferences.DoesNotExist:
        currency = 'USD'  # or any default value you want to set
    #lets create a context that will hold the categories and the expenses
    context = {
        'categories': categories,
        'expenses': expenses,
        'page_obj': page_obj,
        'currency': currency
    }
    return render(request, 'expenses/index.html', context)
# how to add this category first we creat the class category in the models.py then we import to the admin.py then we request useing index and we define the context 
def add_expenses(request):
    categories = Category.objects.all()

    if request.method == 'POST':
        amount = request.POST['amount']
        description = request.POST['description']
        category_id = request.POST['category']
        date = request.POST['date']
        
        if not amount:
            messages.error(request, 'Amount is required')
            context = {
                'categories': categories,
                'values': request.POST
            }
            return render(request, 'expenses/add_expenses.html', context)
        
        category = Category.objects.get(id=category_id)
        user = request.user  # Get the currently logged-in user

        Expense.objects.create(amount=amount, description=description, category=category, date=date, user=user)
        messages.success(request, 'Expense added successfully')
        return redirect('expenses')

    context = {
        'categories': categories
    }
    return render(request, 'expenses/add_expenses.html', context)


#this function to edit the expense
def edit_expense(request, id):
    expense = get_object_or_404(Expense, id=id)
    categories = Category.objects.all()
    if request.method == 'POST':
        try:
            expense.amount = request.POST.get('amount')
            expense.description = request.POST.get('description')
            category_instance = get_object_or_404(Category, id=request.POST.get('category'))
            expense.category = category_instance.name  # Assign the name of the Category as a string
            expense.date = request.POST.get('date')
            expense.save()
            messages.success(request, 'Expense updated successfully')
            return redirect('expenses')
        except Exception as e:
            print(e)
            messages.error(request, f"An error occurred: {e}")
            return redirect('edit-expense', id=id)

    context = {
        'expense': expense,
        'categories': categories
    }
    return render(request, 'expenses/edit-expense.html', context)


def delete_expense(request, id):
    expense = Expense.objects.get(pk=id)
    expense.delete()
    messages.success(request, 'Expense removed')
    return redirect('expenses')

@login_required(login_url='/auth/login')
def expense_category_summary(request):
    today_date = datetime.date.today()
    a_year_ago = today_date - datetime.timedelta(days=365)  # One year ago

    # Print the user making the request
    print(f"Request user: {request.user}")

    # Print the date range
    print(f"Date range: {a_year_ago} to {today_date}")

    # Filter and aggregate expenses by category
    expenses = Expense.objects.filter(date__gte=a_year_ago, date__lte=today_date, user=request.user)
    expense_summary = expenses.values('category').annotate(total_amount=Sum('amount'))

    # Convert the queryset to a dictionary
    finalrep = {expense['category']: expense['total_amount'] for expense in expense_summary}

    # Print the final report dictionary
    print(f"Final report: {finalrep}")

    return JsonResponse({'expense_category_data': finalrep}, safe=False)

#lets create a function that will show the stats of the expenses

def stats_view(request):
    return render(request, 'expenses/stats.html')

def export_csv(request):
    # Create the HttpResponse object with the appropriate CSV header.
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="expenses.csv"'

    writer = csv.writer(response)
    writer.writerow(['Amount', 'Description', 'Category', 'Date'])

    expenses = Expense.objects.filter(user=request.user).values_list('amount', 'description', 'category', 'date')
    for expense in expenses:
        writer.writerow(expense)

    return response


def export_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="expenses.pdf"'

    buffer = []
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleH = styles['Heading1']

    elements = []

    # Title
    title = Paragraph("Expenses Report", styleH)
    elements.append(title)

    # Data for the table
    expenses = Expense.objects.filter(user=request.user)
    data = [["Amount", "Description", "Category", "Date"]]

    for expense in expenses:
        data.append([f"{expense.amount:.2f}", expense.description, expense.category, expense.date.strftime("%Y-%m-%d")])

    # Create the table
    table = Table(data, colWidths=[1.5 * inch, 3 * inch, 2 * inch, 1.5 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))

    elements.append(table)
    doc.build(elements)

    return response

def export_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=expenses.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Expenses'

    columns = ['Amount', 'Description', 'Category', 'Date']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    expenses = Expense.objects.filter(user=request.user)
    for expense in expenses:
        row_num += 1
        worksheet.cell(row=row_num, column=1, value=expense.amount)
        worksheet.cell(row=row_num, column=2, value=expense.description)
        worksheet.cell(row=row_num, column=3, value=expense.category)
        worksheet.cell(row=row_num, column=4, value=expense.date)

    workbook.save(response)
    return response