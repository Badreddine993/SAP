from django.shortcuts import render

# Create your views here.
from django.shortcuts import render ,get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import UserSource, Income
from django.contrib import messages
from django.shortcuts import redirect 
from django.core.paginator import Paginator
import json
from django.http import JsonResponse
from userpreferences.models import UserPreferences
#lets import request form django
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
import openpyxl
import csv
import datetime
from django.db.models import Sum


# Create your views here.
@login_required(login_url='/auth/login')
def search_expenses(request):
    if request.method == 'POST':
        search_str = json.loads(request.body).get('searchText')
        incomes = Income.objects.filter(
            amount__istartswith=search_str, user=request.user
        ) | Income.objects.filter(
            date__istartswith=search_str, user=request.user
        ) | Income.objects.filter(
            description__icontains=search_str, user=request.user
        ) | Income.objects.filter(
            source__name__icontains=search_str, user=request.user  # Corrected to source__name
        )
        data = incomes.values()
        return JsonResponse(list(data), safe=False)

@login_required(login_url='/auth/login')
def index(request):
    sources = UserSource.objects.all()
    incomes = Income.objects.filter(user=request.user)
    paginator = Paginator(incomes, 2)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Safely retrieve UserPreferences
    user_preferences, created = UserPreferences.objects.get_or_create(user=request.user)
    currency = user_preferences.currency if user_preferences.currency else 'USD'  # Default to USD if not set

    context = {
        'sources': sources,
        'incomes': incomes,
        'page_obj': page_obj,
        'currency': currency
    }
    return render(request, 'income/index.html', context)

# how to add this category first we creat the class category in the models.py then we import to the admin.py then we request useing index and we define the context 
def add_Income(request):
    sources = UserSource.objects.all()

    if request.method == 'POST':
        amount = request.POST['amount']
        description = request.POST['description']
        source_id = request.POST['source']
        date = request.POST['date']
        
        if not amount:
            messages.error(request, 'Amount is required')
            context = {
                'sources': sources,
                'values': request.POST
            }
            return render(request, 'income/add_Income.html', context)
        
        source = UserSource.objects.get(id=source_id)
        user = request.user  # Get the currently logged-in user
        Income.objects.create(amount=amount, description=description, source=source, date=date, user=user)
        messages.success(request, 'Income added successfully')
        return redirect('income')

    context = {
        'sources': sources
    }
    return render(request, 'income/add_Income.html', context)


#this function to edit the expense
def edit_Income(request, id):
    income = get_object_or_404(Income, id=id)
    sources = UserSource.objects.all()

    if request.method == 'POST':
        try:
            income.amount = request.POST.get('amount')
            income.description = request.POST.get('description')
            source_instance = get_object_or_404(UserSource, id=request.POST.get('source'))
            income.source = source_instance.name  # Assign the name of the Category as a string
            income.date = request.POST.get('date')
            income.save()
            messages.success(request, 'Income updated successfully')
            return redirect('income')
        except Exception as e:
            print(e)
            messages.error(request, f"An error occurred: {e}")
            return redirect('edit-Income', id=id)

    context = {
        'income': income,
        'sources': sources
    }
    return render(request, 'income/edit-Income.html', context)


def delete_Income(request, id):
    income = Income.objects.get(pk=id)
    income.delete()
    messages.success(request, 'income removed')
    return redirect('income')
#lets create a delete function that will import the expense
def export_income_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="income.csv"'

    writer = csv.writer(response)
    writer.writerow(['Title', 'Amount', 'Date', 'Source', 'Description', 'User'])

    incomes = Income.objects.filter(user=request.user).values_list('title', 'amount', 'date', 'source', 'description', 'user__username')
    for income in incomes:
        writer.writerow(income)

    return response


def export_income_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="income.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    styleN = styles['Normal']
    styleH = styles['Heading1']

    elements = []

    # Title
    title = Paragraph("Income Report", styleH)
    elements.append(title)

    # Data for the table
    incomes = Income.objects.filter(user=request.user)
    data = [["Amount", "Date", "Source", "Description"]]

    for income in incomes:
        data.append([f"{income.amount:.2f}", income.date.strftime("%Y-%m-%d"), income.source, income.description])

    # Create the table
    table = Table(data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 3 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))

    elements.append(table)
    doc.build(elements)

    return response
    

    return response
def export_income_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=income.xlsx'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Income'

    columns = ['Title', 'Amount', 'Date', 'Source', 'Description']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    incomes = Income.objects.filter(user=request.user)
    for income in incomes:
        row_num += 1
        worksheet.cell(row=row_num, column=1, value=income.title)
        worksheet.cell(row=row_num, column=2, value=income.amount)
        worksheet.cell(row=row_num, column=3, value=income.date)
        worksheet.cell(row=row_num, column=4, value=income.source)
        worksheet.cell(row=row_num, column=5, value=income.description)

    workbook.save(response)
    return response



@login_required(login_url='/auth/login')
def income_source_summary(request):
    today_date = datetime.date.today()
    a_year_ago = today_date - datetime.timedelta(days=365)  # One year ago

    # Print the user making the request
    print(f"Request user: {request.user}")

    # Print the date range
    print(f"Date range: {a_year_ago} to {today_date}")

    # Filter and aggregate income by source
    incomes = Income.objects.filter(date__gte=a_year_ago, date__lte=today_date, user=request.user)
    income_summary = incomes.values('source').annotate(total_amount=Sum('amount'))

    # Convert the queryset to a dictionary
    finalrep = {income['source']: income['total_amount'] for income in income_summary}

    # Print the final report dictionary
    print(f"Final report: {finalrep}")

    return JsonResponse({'income_source_data': finalrep}, safe=False)

def stats_view(request):
    return render(request, 'income/stat.html')